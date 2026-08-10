import os
import uuid
import threading
import time
import re
import json
import sqlite3

from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from jusho import Jusho
import posuto

# posuto は日本郵便KEN_ALLの最新データを同梱しており、
# jusho（同梱DBが古い）で見つからない町名のフォールバックに使う
POSUTO_DB = os.path.join(os.path.dirname(posuto.__file__), "postaldata.db")

app = Flask(__name__)
_base_dir = os.environ.get("RENDER", None)
_data_root = "/tmp" if _base_dir is not None else os.path.dirname(__file__)
app.config["UPLOAD_FOLDER"] = os.path.join(_data_root, "uploads")
app.config["RESULT_FOLDER"] = os.path.join(_data_root, "results")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["RESULT_FOLDER"], exist_ok=True)

# Job tracking
jobs = {}

# jusho DB will be created per-thread to avoid SQLite threading issues


WARD_REMAP = {
    "浜松市中央区": ["浜松市中区", "浜松市東区", "浜松市南区"],
    "浜松市浜名区": ["浜松市西区", "浜松市北区", "浜松市浜北区"],
}

# 住所表記で相互に揺れる異体字・仮名
KANJI_ADDR_VARIANTS = {
    "州": "洲", "洲": "州", "磯": "礒", "礒": "磯",
    "槙": "槇", "槇": "槙", "諌": "諫", "諫": "諌",
    "繩": "縄", "縄": "繩", "鴎": "鷗", "鷗": "鴎",
    "一": "壱", "壱": "一",
}

KE_GROUP = ["ヶ", "ケ", "が", "ガ"]  # ヶ ケ が ガ


def _char_variants(s):
    variants = [s]

    # ヶ/ケ/が/ガ は地名で相互に揺れる
    for i, ch in enumerate(s):
        if ch in KE_GROUP:
            expanded = []
            for v in variants:
                for repl in KE_GROUP:
                    expanded.append(v[:i] + repl + v[i + 1:])
            variants = expanded

    # ノ/の の揺れ、および省略
    no_expanded = []
    for v in variants:
        no_expanded.append(v)
        if "ノ" in v or "の" in v:
            no_expanded.append(v.replace("ノ", "の"))
            no_expanded.append(v.replace("の", "ノ"))
            no_expanded.append(v.replace("ノ", "").replace("の", ""))
    variants = no_expanded

    # 異体字
    kanji_expanded = []
    for v in variants:
        kanji_expanded.append(v)
        for old, new in KANJI_ADDR_VARIANTS.items():
            if old in v:
                kanji_expanded.append(v.replace(old, new))
    variants = kanji_expanded

    seen = set()
    result = []
    for v in variants:
        if v and v not in seen:
            seen.add(v)
            result.append(v)
    return result


KANJI_DIGITS = "〇一二三四五六七八九"


def _to_kanji_num(n):
    if n < 10:
        return KANJI_DIGITS[n]
    tens, ones = divmod(n, 10)
    head = "" if tens == 1 else KANJI_DIGITS[tens]
    return head + "十" + (KANJI_DIGITS[ones] if ones else "")


def _jo_town_candidates(rest):
    """北海道式の連番町名（例: 北16条西、屯田7条、東5条南、川端町5条）の候補を作る。

    町名の途中に数字が入るため通常のパースでは数字の直前で切れてしまい、
    「北」だけで検索して別の条にマッチしてしまうのを防ぐ。
    """
    if not rest:
        return []
    a = rest.translate(
        str.maketrans("０１２３４５６７８９", "0123456789")
    )
    m = re.match(r"^([^\d]{1,8}?)(\d{1,2})\s*(条)?\s*([東西南北])?", a)
    if not m:
        return []
    prefix, num, jo, direction = m.group(1), int(m.group(2)), m.group(3), m.group(4)
    if not prefix or num == 0:
        return []
    # 「条」が省略された表記（例: 北25西11）は、方角で始まり方角が続く形のみ対象にする
    if not jo and not (prefix in "東西南北" and direction):
        return []
    out = []
    for num_str in (_to_kanji_num(num), str(num)):
        stem = f"{prefix}{num_str}条"
        if direction:
            out.append(stem + direction)
        out.append(stem)
    return out


def _to_half(s):
    return s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))


def _first_nums(s):
    """先頭から2つの数字を返す（例: 戸山3-19-1 → (3, 19)）。"""
    found = [int(x) for x in re.findall(r"\d+", s)[:2]]
    while len(found) < 2:
        found.append(None)
    return found[0], found[1]


def _is_chome_style(rest):
    """町名の直後の数字が丁目か番地かを推定する。

    同じ町名に丁目用/番地用の2つの郵便番号がある場合の判定に使う。
    丁目は通常 50 以下なので、それを超える数字は番地とみなす。
    """
    if not rest:
        return False
    a = _to_half(rest)
    if "丁目" in a:
        return True
    m = re.search(r"\d+", a)
    return m is not None and int(m.group()) <= 50


def _parse_note_nums(text):
    """「1〜6」「1、2」「18・21」のような注記から数字の集合を作る。"""
    nums = set()
    for part in re.split(r"[、・,，]", _to_half(text)):
        rng = re.match(r"\s*(\d+)\s*[〜~～\-−]\s*(\d+)", part)
        if rng:
            lo, hi = int(rng.group(1)), int(rng.group(2))
            if lo <= hi <= lo + 10000:
                nums.update(range(lo, hi + 1))
            continue
        one = re.search(r"\d+", part)
        if one:
            nums.add(int(one.group()))
    return nums


# 注記と住所の整合度。大きいほど優先して採用する。
NOTE_CONFIRMED = 5   # 注記が住所と明確に一致
NOTE_NONE = 3        # 注記なし
NOTE_SONOTA = 1      # 「その他」= 受け皿
NOTE_MISMATCH = 0    # 注記が住所と明確に不一致


def _note_rank(note, rest, chome_style, num, num2=None):
    """KEN_ALLの注記（丁目/番地/その他/小字名/範囲）と住所の整合度を判定する。

    同じ町名に複数の郵便番号がある場合、どれを採るべきかは注記で決まる。
    例: 新宿区戸山 = 169-0052「3丁目18・21番」/ 162-0052「その他」
    """
    if not note:
        return NOTE_NONE
    if note == "その他":
        return NOTE_SONOTA
    if note == "丁目":
        return NOTE_CONFIRMED if chome_style else NOTE_MISMATCH
    if note == "番地":
        return NOTE_MISMATCH if chome_style else NOTE_CONFIRMED
    # 小字・団地名の列挙（例: 「追分、追分西、上北野、長沼」）
    for part in re.split(r"[、・]", note):
        if part and not re.search(r"\d", part) and part in rest:
            return NOTE_CONFIRMED
    # 丁目の指定（例: 「1〜6丁目」「1、2丁目、3丁目1番〜282番」）
    chome_specs = re.findall(r"([\d０-９、・〜~～\-−]+)丁目", note)
    if chome_specs:
        if not chome_style:
            return NOTE_MISMATCH
        chome_nums = set()
        for spec in chome_specs:
            chome_nums |= _parse_note_nums(spec)
        if num not in chome_nums:
            return NOTE_MISMATCH
        # 丁目の指定が1つだけなら、続く「N番」の限定も判定に使う
        # （例: 戸山 169-0052「3丁目18・21番」）
        if len(chome_specs) == 1 and not _banchi_matches(note, num2):
            return NOTE_MISMATCH
        return NOTE_CONFIRMED
    if "番" in note:
        # 番地の範囲指定（例: 小野町 891-1222「4784〜5118番地」）
        if chome_style or not _banchi_matches(note, num):
            return NOTE_MISMATCH
        return NOTE_CONFIRMED
    # 小字・団地名などの注記が住所に見当たらない場合は「その他」より下げる
    return NOTE_MISMATCH


def _banchi_matches(note, num):
    """注記の「N番」「N番〜M番」「N番以上」等と番地が整合するか判定する。"""
    flat = re.sub(r"番(?=[〜~～\-−])", "", note)
    m = re.search(r"([\d０-９、・〜~～\-−]+)番地?\s*(以上|以下|以内)?", flat)
    if not m:
        return True
    nums = _parse_note_nums(m.group(1))
    if not nums:
        return True
    if num is None:
        return False
    if m.group(2) == "以上":
        return num >= min(nums)
    if m.group(2) in ("以下", "以内"):
        return num <= max(nums)
    return num in nums


def _digit_variants(s):
    half = "0123456789"
    full = "０１２３４５６７８９"
    to_full = s.translate(str.maketrans(half, full))
    to_half = s.translate(str.maketrans(full, half))
    out = [s]
    for v in (to_full, to_half):
        if v not in out:
            out.append(v)
    return out


def _build_city_variants(city):
    variants = [city]
    for new_ward, old_wards in WARD_REMAP.items():
        if city == new_ward:
            variants.extend(old_wards)
            break
    # 郡下の町村が市に昇格したケース（例: 岩手郡滝沢村 → 滝沢市）
    if "郡" in city:
        gm = re.search(r"郡(.+?)[町村]$", city)
        if gm:
            core = gm.group(1)
            variants.extend([core + "市", core + "町", core + "村"])
    # ヶ/ケ 等の揺れ
    expanded = []
    for v in variants:
        expanded.extend(_char_variants(v))
    seen = set()
    out = []
    for v in expanded:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _search_with_city_match(town_clean, pref, city, jusho_db, strict=False,
                            prefer_chome=False, rest="", num=None, num2=None):
    results = jusho_db.search_addresses(town_clean)
    city_variants = _build_city_variants(city)

    exact_match = None
    sonota_match = None
    subarea_match = None
    best_match = None
    for r in results:
        addr_str = str(r)
        if pref not in addr_str:
            continue
        city_matched = False
        for cv in city_variants:
            if cv in addr_str:
                city_matched = True
                break
        if not city_matched:
            city_parts = re.findall(r"[^市区町村郡]+[市区町村]", city)
            if city_parts and all(part in addr_str for part in city_parts):
                city_matched = True
        if not city_matched:
            continue
        zm = re.search(r"〒(\d{3}-\d{4})", addr_str)
        if not zm:
            continue
        zipcode = zm.group(1).replace("-", "")
        # 町名の直後がローマ字開き括弧 "(" = 小字なしの完全一致
        # 町名の直後が末尾 = 完全一致
        if f" {town_clean}(" in addr_str or addr_str.endswith(f" {town_clean}"):
            if exact_match is None:
                exact_match = zipcode
        # 「（その他）」= 小字を特定できない場合の代表郵便番号
        elif f" {town_clean}（その他）" in addr_str:
            if sonota_match is None:
                sonota_match = zipcode
        # 町名の直後が全角括弧 = 小字・丁目範囲付き（その他が無い場合の候補）
        elif f" {town_clean}（" in addr_str:
            if subarea_match is None:
                subarea_match = zipcode
        if best_match is None:
            best_match = zipcode

    if strict:
        return exact_match or sonota_match or subarea_match
    return exact_match or sonota_match or subarea_match or best_match


def _posuto_city_match(town_clean, pref, city, conn, strict=False,
                       prefer_chome=False, rest="", num=None, num2=None):
    rows = conn.execute(
        "SELECT code, city, neighborhood, data FROM postal_data "
        "WHERE prefecture = ? AND neighborhood LIKE ?",
        (pref, "%" + town_clean + "%"),
    ).fetchall()
    city_variants = _build_city_variants(city)

    exact_ranked = []     # 町名完全一致。注記との整合度で選ぶ
    suffix_match = None   # 町名＋「町/村」で一致（例: 礒野東 → 礒野東町）
    subarea_match = None
    best_match = None
    for code, rcity, neighborhood, data in rows:
        city_matched = any(cv in rcity for cv in city_variants)
        if not city_matched:
            city_parts = re.findall(r"[^市区町村郡]+[市区町村]", city)
            if city_parts and all(part in rcity for part in city_parts):
                city_matched = True
        if not city_matched:
            continue
        if neighborhood == town_clean:
            info = json.loads(data)
            # 非分割（partial=false）はその町域＝単一の郵便番号なので信頼度が高い
            exact_ranked.append((
                _note_rank(info.get("note"), rest, prefer_chome, num, num2),
                0 if info.get("partial") else 1,
                code,
            ))
        elif neighborhood in (town_clean + "町", town_clean + "村"):
            if suffix_match is None:
                suffix_match = code
        elif neighborhood.startswith(town_clean):
            if subarea_match is None:
                subarea_match = code
        if best_match is None:
            best_match = code

    # 同じ町名に複数の郵便番号がある場合は注記との整合度が最も高いものを採る
    # （例: 交野市寺 = 576-0063「丁目」/ 576-0006「番地」）
    exact_match = max(exact_ranked)[2] if exact_ranked else None
    if strict:
        return exact_match or suffix_match
    return exact_match or suffix_match or subarea_match or best_match


def _find_zipcode(address, search_fn):
    if not address:
        return ""

    addr = address
    if not re.match(r"^(東京都|北海道|(?:京都|大阪)府|.{2,3}県)", addr):
        city_pref_map = {
            "仙台市": "宮城県", "札幌市": "北海道", "横浜市": "神奈川県",
            "名古屋市": "愛知県", "大阪市": "大阪府", "京都市": "京都府",
            "神戸市": "兵庫県", "福岡市": "福岡県", "広島市": "広島県",
            "さいたま市": "埼玉県", "千葉市": "千葉県", "川崎市": "神奈川県",
            "北九州市": "福岡県", "堺市": "大阪府", "浜松市": "静岡県",
            "新潟市": "新潟県", "熊本市": "熊本県", "岡山市": "岡山県",
            "静岡市": "静岡県", "相模原市": "神奈川県",
        }
        for city, pref in city_pref_map.items():
            if addr.startswith(city):
                addr = pref + addr
                break

    NUM_CHARS = r"[\d０-９]"
    patterns = [
        rf"^(東京都|北海道|(?:京都|大阪)府|.{{2,3}}県)(.+?市.+?区)(.+?)(?:{NUM_CHARS}|$)",
        rf"^(東京都)(.+?区)(.+?)(?:{NUM_CHARS}|$)",
        rf"^(東京都|北海道|(?:京都|大阪)府|.{{2,3}}県)(.+?市)(.+?)(?:{NUM_CHARS}|$)",
        rf"^(東京都|北海道|(?:京都|大阪)府|.{{2,3}}県)(.+?郡.+?(?:町|村))(.+?)(?:{NUM_CHARS}|$)",
        rf"^(東京都|北海道|(?:京都|大阪)府|.{{2,3}}県)(.+?(?:町|村))(.+?)(?:{NUM_CHARS}|$)",
    ]

    for pat in patterns:
        m = re.match(pat, addr)
        if m:
            pref, city, town = m.group(1), m.group(2), m.group(3)
            # 市名に「市」が二つ含まれる誤分割を補正（例: 四日市市/野々市市 → town="市桜町"）
            town = re.sub(r"^市(?=.)", "", town)
            town_clean = re.sub(r"[０-９0-9一二三四五六七八九十丁目番地号の\-－ー・]+$", "", town).strip()
            town_clean = re.sub(r"^(?:大字|字)", "", town_clean)
            if not town_clean and not town:
                continue

            base_towns = []
            # 連番町名（北16条西 等）は数字を含むため最優先で試す
            rest = addr[m.end(2):]
            jo_candidates = _jo_town_candidates(rest)
            base_towns.extend(jo_candidates)
            if town_clean:
                base_towns.append(town_clean)
            # 元の町名（末尾の丁目等を除去する前）も試す（例: 六番丁）
            raw_town = re.sub(r"^(?:大字|字)", "", town)
            if raw_town and raw_town not in base_towns:
                base_towns.append(raw_town)
            # 市名が町名に重複して残るケースを除去（例: 富谷市富谷町成田 → 成田）
            city_core = re.sub(r"^.+郡", "", city)
            city_core = re.sub(r"[市区町村]$", "", city_core)
            if city_core and town_clean.startswith(city_core):
                deduped = re.sub(r"^[市区町村]", "", town_clean[len(city_core):])
                if deduped and deduped != town_clean:
                    base_towns.append(deduped)
            if not town_clean:
                town_clean = raw_town
            # 「字」「大字」の処理: 除去 / 分割（例: 脇町大字脇町 → 脇町脇町、滝沢字牧野林 → 牧野林）
            if "字" in town_clean:
                base_towns.append(town_clean.replace("大字", "").replace("字", ""))
                aza_split = [p for p in re.split(r"大字|字", town_clean) if p]
                # 先頭からの累積結合を長い順に試す
                # （例: 一箕町/八幡/八幡 → 一箕町八幡八幡、一箕町八幡、一箕町）
                for k in range(len(aza_split), 0, -1):
                    base_towns.append("".join(aza_split[:k]))
                base_towns.extend(aza_split)
            # 京都の通り名住所（例: 壬生通八条下ル東寺町 → 東寺町）
            kyoto_split = re.split(r"(?:上る|下る|上ル|下ル|東入る|西入る|東入ル|西入ル|東入|西入)", town_clean)
            if len(kyoto_split) > 1 and kyoto_split[-1]:
                base_towns.append(kyoto_split[-1])
            # 複合地名のフォールバック: 末尾から1文字ずつ短縮（例: 西今宿阿弥陀寺 → 西今宿）
            # 「町」境界での区切り（例: 明大寺町伝馬 → 明大寺町）もこの短縮で網羅される
            for cut in range(len(town_clean) - 1, 1, -1):
                base_towns.append(town_clean[:cut])
            # 末尾「町」の除去（例: 佐藤町 → 佐藤）
            if town_clean.endswith("町") and len(town_clean) > 2:
                base_towns.append(town_clean[:-1])
            # 最終手段: 先頭の余分な1〜2文字を除去（例: 元データ誤記「立柏の森」→「柏の森」）
            for drop in (1, 2):
                if len(town_clean) - drop >= 2:
                    base_towns.append(town_clean[drop:])

            # 町名の直後の数字が丁目とみなせる大きさなら丁目表記と判定する。
            # （例: 寺3-20-1 → 3丁目 / 越ヶ谷2788-1 → 番地）
            # 短縮して得た候補（小字を落とした後）は丁目表記とはしない
            rest_half = _to_half(rest)
            first_num, first_num2 = _first_nums(rest_half)
            chome_style = _is_chome_style(rest)
            jo_set = set(jo_candidates)
            # 連番町名では条の数字の後ろに丁目が来る（例: 北16条西2 → 2丁目）
            jo_m = re.match(r"^[^\d]{1,8}?\d{1,2}\s*条?\s*[東西南北]?", rest_half)
            jo_num, jo_num2 = (_first_nums(rest_half[jo_m.end():]) if jo_m
                               else (None, None))

            # 1周目は町名が完全一致する候補のみ採用し、2周目で前方一致等を許容する。
            # 短縮した町名が別町域に前方一致してしまうのを防ぐ（例: 長良小山田 →
            # 「長良小」が長良小松町に一致するより先に「長良」の完全一致を採る）
            for strict in (True, False):
                seen = set()
                for base in base_towns:
                    is_jo = base in jo_set
                    prefer_chome = is_jo or (
                        chome_style and base in (town_clean, raw_town)
                    )
                    num = jo_num if is_jo else first_num
                    num2 = jo_num2 if is_jo else first_num2
                    for tv in _char_variants(base):
                        for dv in _digit_variants(tv):
                            if dv in seen or not dv:
                                continue
                            seen.add(dv)
                            result = search_fn(dv, pref, city, strict,
                                               prefer_chome, rest, num, num2)
                            if result:
                                return result
    return ""


def address_to_zipcode(address, jusho_db, posuto_conn=None):
    # 常に最新の日本郵便KEN_ALL（posuto）を主データ源として優先
    if posuto_conn is not None:
        result = _find_zipcode(
            address,
            lambda dv, pref, city, strict, prefer_chome, rest, num, num2:
                _posuto_city_match(
                    dv, pref, city, posuto_conn, strict, prefer_chome,
                    rest, num, num2
                ),
        )
        if result:
            return result
    # posutoで見つからない場合のみ jusho（旧DB）にフォールバック
    return _find_zipcode(
        address,
        lambda dv, pref, city, strict, prefer_chome, rest, num, num2:
            _search_with_city_match(
                dv, pref, city, jusho_db, strict, prefer_chome,
                rest, num, num2
            ),
    )


def process_excel(job_id, input_path, output_path):
    try:
        jobs[job_id]["status"] = "processing"
        # Create per-thread jusho instance to avoid SQLite threading issues
        thread_jusho = Jusho()
        thread_posuto = sqlite3.connect(POSUTO_DB, check_same_thread=False)

        wb = openpyxl.load_workbook(input_path)
        ws = wb.active

        ws["B1"] = "高校名"
        ws["C1"] = "郵便番号（元データ）"
        ws["D1"] = "郵便番号（住所から逆引き）"
        ws["E1"] = "住所（元データ）"

        total = ws.max_row - 1
        jobs[job_id]["total"] = total

        for row_idx in range(2, ws.max_row + 1):
            school_name = ws.cell(row=row_idx, column=2).value
            zipcode = ws.cell(row=row_idx, column=3).value
            original_address = ws.cell(row=row_idx, column=4).value

            reverse_zipcode = address_to_zipcode(original_address, thread_jusho, thread_posuto)

            ws.cell(row=row_idx, column=4).value = reverse_zipcode
            ws.cell(row=row_idx, column=5).value = original_address

            jobs[job_id]["progress"] = row_idx - 1
            jobs[job_id]["current_school"] = school_name or ""

        wb.save(output_path)
        jobs[job_id]["status"] = "done"
        jobs[job_id]["progress"] = total

    except Exception as e:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"] = str(e)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "ファイルが選択されていません"}), 400
    f = request.files["file"]
    if not f.filename or not f.filename.endswith(".xlsx"):
        return jsonify({"error": ".xlsx ファイルを選択してください"}), 400

    job_id = str(uuid.uuid4())[:8]
    input_path = os.path.join(app.config["UPLOAD_FOLDER"], f"{job_id}_input.xlsx")
    output_path = os.path.join(app.config["RESULT_FOLDER"], f"{job_id}_output.xlsx")
    f.save(input_path)

    jobs[job_id] = {
        "status": "queued",
        "progress": 0,
        "total": 0,
        "current_school": "",
        "output_path": output_path,
        "original_filename": f.filename,
    }

    thread = threading.Thread(target=process_excel, args=(job_id, input_path, output_path))
    thread.daemon = True
    thread.start()

    return jsonify({"job_id": job_id})


@app.route("/status/<job_id>")
def status(job_id):
    if job_id not in jobs:
        return jsonify({"error": "ジョブが見つかりません"}), 404
    job = jobs[job_id]
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "current_school": job.get("current_school", ""),
        "error": job.get("error", ""),
    })


@app.route("/download/<job_id>")
def download(job_id):
    # jobs はプロセス内メモリのため、ワーカー再起動などで失われることがある。
    # その場合でも結果ファイルがディスクに残っていれば配信できるようにする。
    default_path = os.path.join(
        app.config["RESULT_FOLDER"], f"{job_id}_output.xlsx"
    )
    job = jobs.get(job_id)

    if job is not None:
        if job["status"] != "done":
            return jsonify({"error": "処理がまだ完了していません"}), 400
        output_path = job["output_path"]
        original = job.get("original_filename", "output.xlsx")
    else:
        if not os.path.exists(default_path):
            return jsonify({
                "error": "時間が経過したため処理結果が失われました。"
                         "お手数ですが、もう一度アップロードして処理してください。"
            }), 404
        output_path = default_path
        original = "output.xlsx"

    name_base = os.path.splitext(original)[0]
    download_name = f"{name_base}_processed.xlsx"

    return send_file(output_path, as_attachment=True, download_name=download_name)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
